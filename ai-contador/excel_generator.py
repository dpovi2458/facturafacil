"""
Generador de reportes Excel profesionales
Crea reportes contables con formato empresarial
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, REPORT_CONFIG


# Estilos
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
CURRENCY_FORMAT = '"S/ "#,##0.00'
DATE_FORMAT = 'DD/MM/YYYY'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def apply_header_style(cell):
    """Aplica estilo de encabezado a una celda"""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER


def apply_currency_format(ws, col, start_row, end_row):
    """Aplica formato de moneda a una columna"""
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT


def auto_adjust_columns(ws):
    """Ajusta automáticamente el ancho de las columnas"""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        length = min(length + 2, 50)  # Máximo 50 caracteres
        ws.column_dimensions[column_cells[0].column_letter].width = length


def generate_sales_report(
    business_info: dict,
    documents: pd.DataFrame,
    sales_summary: pd.DataFrame,
    top_clients: pd.DataFrame,
    top_products: pd.DataFrame,
    ai_analysis: dict,
    filename: str = None
) -> str:
    """
    Genera reporte completo de ventas en Excel
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_ventas_{timestamp}.xlsx"
    
    filepath = REPORTS_DIR / filename
    wb = Workbook()
    
    # =========================================
    # HOJA 1: RESUMEN EJECUTIVO
    # =========================================
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"
    
    # Título
    ws_resumen.merge_cells('A1:F1')
    title_cell = ws_resumen['A1']
    title_cell.value = f"📊 REPORTE DE VENTAS - {business_info.get('razon_social', 'Mi Negocio')}"
    title_cell.font = Font(bold=True, size=16, color="1E40AF")
    title_cell.alignment = Alignment(horizontal='center')
    
    # Información del negocio
    ws_resumen['A3'] = "RUC:"
    ws_resumen['B3'] = business_info.get('ruc', '')
    ws_resumen['A4'] = "Razón Social:"
    ws_resumen['B4'] = business_info.get('razon_social', '')
    ws_resumen['A5'] = "Fecha de Reporte:"
    ws_resumen['B5'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    # Métricas principales
    row = 8
    ws_resumen[f'A{row}'] = "📈 MÉTRICAS PRINCIPALES"
    ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
    
    if ai_analysis and 'resumen' in ai_analysis:
        resumen = ai_analysis['resumen']
        row += 2
        metrics = [
            ("Total Ventas", f"S/ {resumen.get('total_ventas', 0):,.2f}"),
            ("Promedio Mensual", f"S/ {resumen.get('promedio_mensual', 0):,.2f}"),
            ("Total Documentos", resumen.get('total_documentos', 0)),
            ("IGV Acumulado", f"S/ {resumen.get('igv_total', 0):,.2f}"),
            ("Tendencia", resumen.get('tendencia', 'N/A'))
        ]
        for metric, value in metrics:
            ws_resumen[f'A{row}'] = metric
            ws_resumen[f'B{row}'] = value
            ws_resumen[f'A{row}'].font = Font(bold=True)
            row += 1
    
    # Insights de IA
    row += 2
    ws_resumen[f'A{row}'] = "🤖 ANÁLISIS INTELIGENTE"
    ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    if ai_analysis:
        for insight in ai_analysis.get('insights', []):
            ws_resumen[f'A{row}'] = f"• {insight}"
            row += 1
        
        row += 1
        ws_resumen[f'A{row}'] = "💡 RECOMENDACIONES"
        ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for rec in ai_analysis.get('recomendaciones', []):
            ws_resumen[f'A{row}'] = f"• {rec}"
            row += 1
    
    auto_adjust_columns(ws_resumen)
    
    # =========================================
    # HOJA 2: DETALLE DE DOCUMENTOS
    # =========================================
    ws_docs = wb.create_sheet("Documentos")
    
    if not documents.empty:
        # Encabezados
        headers = ['Tipo', 'Serie-Número', 'Fecha', 'Cliente', 'Subtotal', 'IGV', 'Total', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws_docs.cell(row=1, column=col, value=header)
            apply_header_style(cell)
        
        # Datos
        for row_idx, doc in documents.iterrows():
            row = row_idx + 2
            ws_docs.cell(row=row, column=1, value=doc.get('tipo', '').upper())
            ws_docs.cell(row=row, column=2, value=f"{doc.get('serie', '')}-{doc.get('numero', '')}")
            ws_docs.cell(row=row, column=3, value=doc.get('fecha_emision', ''))
            ws_docs.cell(row=row, column=4, value=doc.get('cliente_nombre', 'Cliente General'))
            ws_docs.cell(row=row, column=5, value=doc.get('subtotal', 0))
            ws_docs.cell(row=row, column=6, value=doc.get('igv', 0))
            ws_docs.cell(row=row, column=7, value=doc.get('total', 0))
            ws_docs.cell(row=row, column=8, value=doc.get('estado', '').upper())
        
        # Formato de moneda
        last_row = len(documents) + 1
        for col in [5, 6, 7]:
            apply_currency_format(ws_docs, col, 2, last_row)
        
        # Totales
        total_row = last_row + 2
        ws_docs.cell(row=total_row, column=4, value="TOTALES:")
        ws_docs.cell(row=total_row, column=4).font = Font(bold=True)
        ws_docs.cell(row=total_row, column=5, value=documents['subtotal'].sum())
        ws_docs.cell(row=total_row, column=6, value=documents['igv'].sum())
        ws_docs.cell(row=total_row, column=7, value=documents['total'].sum())
        for col in [5, 6, 7]:
            ws_docs.cell(row=total_row, column=col).number_format = CURRENCY_FORMAT
            ws_docs.cell(row=total_row, column=col).font = Font(bold=True)
    
    auto_adjust_columns(ws_docs)
    
    # =========================================
    # HOJA 3: RESUMEN MENSUAL
    # =========================================
    ws_mensual = wb.create_sheet("Resumen Mensual")
    
    if not sales_summary.empty:
        headers = ['Año', 'Mes', 'Tipo', 'Documentos', 'Subtotal', 'IGV', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws_mensual.cell(row=1, column=col, value=header)
            apply_header_style(cell)
        
        for row_idx, data in sales_summary.iterrows():
            row = row_idx + 2
            ws_mensual.cell(row=row, column=1, value=data.get('año', ''))
            ws_mensual.cell(row=row, column=2, value=data.get('mes', ''))
            ws_mensual.cell(row=row, column=3, value=data.get('tipo', '').upper())
            ws_mensual.cell(row=row, column=4, value=data.get('cantidad_documentos', 0))
            ws_mensual.cell(row=row, column=5, value=data.get('subtotal', 0))
            ws_mensual.cell(row=row, column=6, value=data.get('igv', 0))
            ws_mensual.cell(row=row, column=7, value=data.get('total', 0))
        
        last_row = len(sales_summary) + 1
        for col in [5, 6, 7]:
            apply_currency_format(ws_mensual, col, 2, last_row)
        
        # Gráfico de barras
        if len(sales_summary) > 1:
            chart = BarChart()
            chart.title = "Ventas por Mes"
            chart.x_axis.title = "Período"
            chart.y_axis.title = "Total (S/)"
            data_ref = Reference(ws_mensual, min_col=7, min_row=1, max_row=last_row)
            cats_ref = Reference(ws_mensual, min_col=2, min_row=2, max_row=last_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 15
            chart.height = 10
            ws_mensual.add_chart(chart, "I2")
    
    auto_adjust_columns(ws_mensual)
    
    # =========================================
    # HOJA 4: TOP CLIENTES
    # =========================================
    ws_clientes = wb.create_sheet("Top Clientes")
    
    if not top_clients.empty:
        headers = ['Cliente', 'Documento', 'Total Compras', 'Monto Total', 'Última Compra']
        for col, header in enumerate(headers, 1):
            cell = ws_clientes.cell(row=1, column=col, value=header)
            apply_header_style(cell)
        
        for row_idx, client in top_clients.iterrows():
            row = row_idx + 2
            ws_clientes.cell(row=row, column=1, value=client.get('nombre', ''))
            ws_clientes.cell(row=row, column=2, value=client.get('numero_documento', ''))
            ws_clientes.cell(row=row, column=3, value=client.get('total_compras', 0))
            ws_clientes.cell(row=row, column=4, value=client.get('monto_total', 0))
            ws_clientes.cell(row=row, column=5, value=client.get('ultima_compra', ''))
        
        apply_currency_format(ws_clientes, 4, 2, len(top_clients) + 1)
    
    auto_adjust_columns(ws_clientes)
    
    # =========================================
    # HOJA 5: TOP PRODUCTOS
    # =========================================
    ws_productos = wb.create_sheet("Top Productos")
    
    if not top_products.empty:
        headers = ['Producto', 'Cantidad Vendida', 'Monto Total', 'En Documentos']
        for col, header in enumerate(headers, 1):
            cell = ws_productos.cell(row=1, column=col, value=header)
            apply_header_style(cell)
        
        for row_idx, prod in top_products.iterrows():
            row = row_idx + 2
            ws_productos.cell(row=row, column=1, value=prod.get('descripcion', ''))
            ws_productos.cell(row=row, column=2, value=prod.get('cantidad_vendida', 0))
            ws_productos.cell(row=row, column=3, value=prod.get('monto_total', 0))
            ws_productos.cell(row=row, column=4, value=prod.get('en_documentos', 0))
        
        apply_currency_format(ws_productos, 3, 2, len(top_products) + 1)
        
        # Gráfico de pastel
        if len(top_products) > 1:
            chart = PieChart()
            chart.title = "Productos Más Vendidos"
            data_ref = Reference(ws_productos, min_col=3, min_row=1, max_row=min(len(top_products) + 1, 6))
            cats_ref = Reference(ws_productos, min_col=1, min_row=2, max_row=min(len(top_products) + 1, 6))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 12
            chart.height = 10
            ws_productos.add_chart(chart, "F2")
    
    auto_adjust_columns(ws_productos)
    
    # Guardar
    wb.save(filepath)
    return str(filepath)


def generate_tax_report(
    business_info: dict,
    documents: pd.DataFrame,
    year: int,
    month: int,
    filename: str = None
) -> str:
    """
    Genera reporte tributario mensual para declaración SUNAT
    """
    if filename is None:
        filename = f"reporte_tributario_{year}_{month:02d}.xlsx"
    
    filepath = REPORTS_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Declaración Mensual"
    
    # Filtrar documentos del mes
    docs_mes = documents[
        (documents['fecha_emision'].str[:7] == f"{year}-{month:02d}")
    ] if not documents.empty else pd.DataFrame()
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"REPORTE TRIBUTARIO - {month:02d}/{year}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Info del negocio
    ws['A3'] = "RUC:"
    ws['B3'] = business_info.get('ruc', '')
    ws['A4'] = "Razón Social:"
    ws['B4'] = business_info.get('razon_social', '')
    
    # Resumen de ventas
    row = 7
    ws[f'A{row}'] = "RESUMEN DE VENTAS DEL MES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 2
    if not docs_mes.empty:
        # Boletas
        boletas = docs_mes[docs_mes['tipo'] == 'boleta']
        ws[f'A{row}'] = "BOLETAS DE VENTA"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "Cantidad:"
        ws[f'B{row}'] = len(boletas)
        row += 1
        ws[f'A{row}'] = "Base Imponible:"
        ws[f'B{row}'] = boletas['subtotal'].sum() if not boletas.empty else 0
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        row += 1
        ws[f'A{row}'] = "IGV:"
        ws[f'B{row}'] = boletas['igv'].sum() if not boletas.empty else 0
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        row += 1
        ws[f'A{row}'] = "Total:"
        ws[f'B{row}'] = boletas['total'].sum() if not boletas.empty else 0
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 2
        # Facturas
        facturas = docs_mes[docs_mes['tipo'] == 'factura']
        ws[f'A{row}'] = "FACTURAS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "Cantidad:"
        ws[f'B{row}'] = len(facturas)
        row += 1
        ws[f'A{row}'] = "Base Imponible:"
        ws[f'B{row}'] = facturas['subtotal'].sum() if not facturas.empty else 0
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        row += 1
        ws[f'A{row}'] = "IGV:"
        ws[f'B{row}'] = facturas['igv'].sum() if not facturas.empty else 0
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        row += 1
        ws[f'A{row}'] = "Total:"
        ws[f'B{row}'] = facturas['total'].sum() if not facturas.empty else 0
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        ws[f'B{row}'].font = Font(bold=True)
        
        # Totales generales
        row += 3
        ws[f'A{row}'] = "TOTALES PARA DECLARACIÓN"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1E40AF")
        row += 1
        ws[f'A{row}'] = "Base Imponible Total:"
        ws[f'B{row}'] = docs_mes['subtotal'].sum()
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "IGV por Pagar:"
        ws[f'B{row}'] = docs_mes['igv'].sum()
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        ws[f'B{row}'].font = Font(bold=True, color="FF0000")
        row += 1
        ws[f'A{row}'] = "Total Ventas:"
        ws[f'B{row}'] = docs_mes['total'].sum()
        ws[f'B{row}'].number_format = CURRENCY_FORMAT
        ws[f'B{row}'].font = Font(bold=True)
    else:
        ws[f'A{row}'] = "No hay documentos emitidos en este período"
    
    auto_adjust_columns(ws)
    wb.save(filepath)
    return str(filepath)
